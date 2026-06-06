"""
Super OPL API — DMC VCCU PM Assistant
Unified Open Point List with full CRUD, import/export, and team member support.

Storage: outputs/super_opl.json  (append-only JSON array, soft-delete)

Endpoints:
  GET    /api/opl                  list with filters + pagination
  POST   /api/opl                  create entry
  GET    /api/opl/{id}             single entry + subtasks
  PUT    /api/opl/{id}             update entry
  DELETE /api/opl/{id}             soft delete (status=deleted)
  PUT    /api/opl/{id}/status      quick status change
  POST   /api/opl/import           bulk import from Excel/CSV
  GET    /api/opl/export           export to Excel
  GET    /api/opl/team-members     team members from active project config
"""

import json
import logging
import os
import io
import tempfile
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Optional, List

from fastapi import APIRouter, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

logger = logging.getLogger("opl_standalone.super_opl")
router = APIRouter()

_ROOT       = Path(__file__).resolve().parent.parent
_OUTPUT_DIR = Path(os.getenv("OUTPUT_DIR", str(_ROOT / "outputs")))
_OPL_FILE   = _OUTPUT_DIR / "super_opl.json"

# ── ID generator ──────────────────────────────────────────────────────────────

def _next_id(entries: list) -> str:
    nums = []
    for e in entries:
        eid = e.get("id", "")
        # Only count OPL-NNN local IDs, ignore BOSCH-xxx IDs
        if eid.startswith("OPL-"):
            try:
                nums.append(int(eid.split("-")[1]))
            except Exception:
                pass
    n = max(nums) + 1 if nums else 1
    return f"OPL-{n:03d}"


# ── Persistence ───────────────────────────────────────────────────────────────

def _load() -> list:
    _OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if not _OPL_FILE.exists():
        return []
    try:
        return json.loads(_OPL_FILE.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        logger.error("super_opl.json is corrupt: %s", e)
        return []
    except OSError as e:
        logger.error("Cannot read super_opl.json: %s", e)
        return []


def _save(entries: list):
    """Atomic write: write to temp file then rename to avoid partial writes."""
    _OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    data = json.dumps(entries, indent=2, ensure_ascii=False)
    try:
        fd, tmp_path = tempfile.mkstemp(dir=_OUTPUT_DIR, suffix=".tmp")
        try:
            with os.fdopen(fd, "w", encoding="utf-8") as f:
                f.write(data)
            os.replace(tmp_path, _OPL_FILE)
        except Exception:
            os.unlink(tmp_path)
            raise
    except OSError as e:
        logger.error("Failed to save super_opl.json: %s", e)
        raise


# ── Pydantic models ───────────────────────────────────────────────────────────

class OPLEntry(BaseModel):
    entry_type:    str = "Task"       # Task | Information | Decision | Subtask
    parent_id:     Optional[str] = None
    subject:       str = ""
    description:   str = ""
    remarks:       str = ""
    last_note:     str = ""
    owner:         str = ""
    responsible:   List[str] = []
    information_to: List[str] = []
    topic_owner:   str = ""
    sw_category:   str = ""
    topic:         str = ""
    sub_topic:     str = ""
    category:      str = ""
    source:        str = ""
    register:      str = ""
    tags:          List[str] = []
    sprint_tag:    str = ""
    meeting:       str = ""
    priority:      str = "Medium"    # High | Medium | Low
    status:        str = "running"   # running|closed|overdue|rejected|waiting_approval|draft|on_hold|deleted
    risk_flag:     bool = False
    risk_impact:   str = ""
    confidential:  bool = False
    pinned:         bool = False
    linked_risk_id: Optional[str] = None   # risk ID this task is a measure for (e.g. LRISK-001, BRISK-376566)
    input_date:     Optional[str] = None
    start_date:     Optional[str] = None
    due_date:       Optional[str] = None
    closed_date:    Optional[str] = None
    input_by:       str = ""
    notes:          List[dict] = []
    attachments:    List[dict] = []


class OPLUpdate(BaseModel):
    """Allowed fields for PUT /api/opl/{id}. Immutable fields (id, created_at, input_date, input_by) are excluded."""
    entry_type:     Optional[str] = None
    parent_id:      Optional[str] = None
    subject:        Optional[str] = None
    description:    Optional[str] = None
    remarks:        Optional[str] = None
    last_note:      Optional[str] = None
    owner:          Optional[str] = None
    responsible:    Optional[List[str]] = None
    information_to: Optional[List[str]] = None
    topic_owner:    Optional[str] = None
    sw_category:    Optional[str] = None
    topic:          Optional[str] = None
    sub_topic:      Optional[str] = None
    category:       Optional[str] = None
    source:         Optional[str] = None
    register:       Optional[str] = None
    tags:           Optional[List[str]] = None
    sprint_tag:     Optional[str] = None
    meeting:        Optional[str] = None
    priority:       Optional[str] = None
    status:         Optional[str] = None
    risk_flag:      Optional[bool] = None
    risk_impact:    Optional[str] = None
    confidential:   Optional[bool] = None
    pinned:         Optional[bool] = None
    linked_risk_id: Optional[str] = None
    start_date:     Optional[str] = None
    due_date:       Optional[str] = None
    closed_date:    Optional[str] = None
    last_change_by: Optional[str] = None
    notes:          Optional[List[dict]] = None
    attachments:    Optional[List[dict]] = None
    bosch_task_id:  Optional[str] = None


class StatusUpdate(BaseModel):
    status: str
    note:   Optional[str] = None
    by:     Optional[str] = None


# ── Helpers ───────────────────────────────────────────────────────────────────

def _now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _today() -> str:
    return date.today().isoformat()


def _is_overdue(entry: dict) -> bool:
    due = entry.get("due_date")
    if not due:
        return False
    status = entry.get("status", "")
    if status in ("closed", "deleted", "rejected"):
        return False
    try:
        return date.fromisoformat(due) < date.today()
    except Exception:
        return False


def _apply_overdue(entry: dict) -> dict:
    """Auto-mark overdue status on read."""
    if _is_overdue(entry) and entry.get("status") == "running":
        entry = dict(entry, status="overdue")
    return entry


# ── STATUS FILTER GROUPS ──────────────────────────────────────────────────────
# Maps "show" filter value → set of statuses to include (None = all)
_SHOW_FILTERS = {
    "running":                    {"running", "overdue"},
    "closed":                     {"closed"},
    "overdue":                    {"overdue"},
    "rejected":                   {"rejected"},
    "waiting_approval":           {"waiting_approval"},
    "draft":                      {"draft"},
    "on_hold":                    {"on_hold"},
    "running_decisions":          {"running", "overdue"},          # + entry_type=Decision
    "running_decisions_info":     {"running", "overdue"},          # + entry_type in (Decision,Information)
    "all_deleted":                None,                            # truly all including deleted
    "all":                        {"running","overdue","closed","rejected","waiting_approval","draft","on_hold"},
}


# ── GET /api/opl ──────────────────────────────────────────────────────────────

@router.get("/api/opl")
def list_opl(
    show:     str = Query("all",  description="Status filter group"),
    category: str = Query("",    description="Filter by category"),
    owner:    str = Query("",    description="Filter by owner"),
    search:   str = Query("",    description="Full-text search on subject/description"),
    tags:     str = Query("",    description="Comma-separated tag filter"),
    priority: str = Query("",    description="Filter by priority"),
    register: str = Query("",    description="Filter by register"),
    page:     int = Query(1,     ge=1),
    limit:    int = Query(100,   ge=1, le=500),
):
    allowed = _SHOW_FILTERS.get(show)
    decision_only = show == "running_decisions"
    decision_info_only = show == "running_decisions_info"
    owner_q   = owner.lower() if owner else ""
    category_q = category.lower() if category else ""
    priority_q = priority.lower() if priority else ""
    register_q = register.lower() if register else ""
    tag_list  = [t.strip().lower() for t in tags.split(",") if t.strip()] if tags else []
    search_q  = search.lower() if search else ""

    def _matches(e: dict) -> bool:
        e = _apply_overdue(e)
        status = e.get("status", "running")
        if allowed is not None and status not in allowed:
            return False
        et = e.get("entry_type", "")
        if decision_only and et != "Decision":
            return False
        if decision_info_only and et not in ("Decision", "Information"):
            return False
        if category_q and e.get("category", "").lower() != category_q:
            return False
        if owner_q and e.get("owner", "").lower() != owner_q and owner_q not in [r.lower() for r in e.get("responsible", [])]:
            return False
        if priority_q and e.get("priority", "").lower() != priority_q:
            return False
        if register_q and e.get("register", "").lower() != register_q:
            return False
        if tag_list and not any(t in [x.lower() for x in e.get("tags", [])] for t in tag_list):
            return False
        if search_q and not (
            search_q in e.get("subject", "").lower() or
            search_q in e.get("description", "").lower() or
            search_q in e.get("remarks", "").lower() or
            search_q in e.get("last_note", "").lower()
        ):
            return False
        return True

    entries = [_apply_overdue(e) for e in _load() if _matches(e)]

    # Pinned items first, then by input_date desc
    entries.sort(key=lambda e: (not e.get("pinned", False), e.get("input_date", "") or ""), reverse=False)
    entries.sort(key=lambda e: e.get("pinned", False), reverse=True)

    total = len(entries)
    start = (page - 1) * limit
    page_entries = entries[start: start + limit]

    return {"total": total, "page": page, "limit": limit, "items": page_entries}


# ── POST /api/opl ─────────────────────────────────────────────────────────────

@router.post("/api/opl", status_code=201)
def create_opl(entry: OPLEntry):
    entries = _load()
    now = _now_iso()
    today = _today()
    new_entry = entry.model_dump()
    new_entry["id"]            = _next_id(entries)
    new_entry["input_date"]    = new_entry.get("input_date") or today
    new_entry["start_date"]    = new_entry.get("start_date") or today
    new_entry["due_date"]      = new_entry.get("due_date") or (date.today() + timedelta(days=14)).isoformat()
    new_entry["last_change"]   = now
    new_entry["last_change_by"] = new_entry.get("input_by") or new_entry.get("owner") or ""
    new_entry["created_at"]    = now
    entries.append(new_entry)
    _save(entries)
    return new_entry


# ── GET /api/opl/{id} ─────────────────────────────────────────────────────────

@router.get("/api/opl/{entry_id}")
def get_opl(entry_id: str):
    entries = _load()
    entry = next((e for e in entries if e["id"] == entry_id), None)
    if not entry:
        raise HTTPException(status_code=404, detail=f"OPL entry '{entry_id}' not found")
    entry = _apply_overdue(entry)
    subtasks = [_apply_overdue(e) for e in entries if e.get("parent_id") == entry_id]
    return {**entry, "subtasks": subtasks}


# ── PUT /api/opl/{id} ─────────────────────────────────────────────────────────

@router.put("/api/opl/{entry_id}")
def update_opl(entry_id: str, updates: OPLUpdate):
    entries = _load()
    idx = next((i for i, e in enumerate(entries) if e["id"] == entry_id), None)
    if idx is None:
        raise HTTPException(status_code=404, detail=f"OPL entry '{entry_id}' not found")
    # Only apply fields that were explicitly set (not None)
    for k, v in updates.model_dump(exclude_none=True).items():
        entries[idx][k] = v
    entries[idx]["last_change"]    = _now_iso()
    entries[idx]["last_change_by"] = updates.last_change_by or entries[idx].get("last_change_by", "")
    _save(entries)
    logger.info("Updated OPL entry %s: fields=%s", entry_id, list(updates.model_dump(exclude_none=True).keys()))
    return entries[idx]


# ── DELETE /api/opl/{id} — soft delete ───────────────────────────────────────

@router.delete("/api/opl/{entry_id}")
def delete_opl(entry_id: str):
    entries = _load()
    idx = next((i for i, e in enumerate(entries) if e["id"] == entry_id), None)
    if idx is None:
        raise HTTPException(status_code=404, detail=f"OPL entry '{entry_id}' not found")
    entries[idx]["status"]        = "deleted"
    entries[idx]["last_change"]   = _now_iso()
    _save(entries)
    return {"status": "ok", "id": entry_id}


# ── PUT /api/opl/{id}/status ──────────────────────────────────────────────────

@router.put("/api/opl/{entry_id}/status")
def update_status(entry_id: str, body: StatusUpdate):
    valid = {"running","closed","overdue","rejected","waiting_approval","draft","on_hold","deleted"}
    if body.status not in valid:
        raise HTTPException(status_code=400, detail=f"Invalid status '{body.status}'")
    entries = _load()
    idx = next((i for i, e in enumerate(entries) if e["id"] == entry_id), None)
    if idx is None:
        raise HTTPException(status_code=404, detail=f"OPL entry '{entry_id}' not found")
    entries[idx]["status"]      = body.status
    entries[idx]["last_change"] = _now_iso()
    if body.status == "closed":
        entries[idx]["closed_date"] = _today()
    if body.note:
        entries[idx].setdefault("notes", []).append({
            "text": body.note,
            "by":   body.by or "",
            "at":   _now_iso(),
        })
        entries[idx]["last_note"] = body.note
    if body.by:
        entries[idx]["last_change_by"] = body.by
    _save(entries)
    return entries[idx]


# ── GET /api/opl/team-members ─────────────────────────────────────────────────

@router.get("/api/opl/team-members")
def get_team_members():
    try:
        from backend.project_config import get_active_project
        cfg = get_active_project()
        members = cfg.get("team_members", [])
    except Exception:
        members = []
    return {"members": members}


# ── POST /api/opl/import ─────────────────────────────────────────────────────

@router.post("/api/opl/import")
async def import_opl(
    file: UploadFile = File(...),
    mode: str = Query("append", description="append | upsert"),
):
    """Accept .xlsx or .csv, map columns, return preview + import result."""
    content = await file.read()
    filename = file.filename or ""
    rows = []

    # ── Parse file ────────────────────────────────────────────────────────────
    if filename.lower().endswith(".csv"):
        import csv, codecs
        reader = csv.DictReader(codecs.iterdecode(io.BytesIO(content), "utf-8-sig"))
        rows = [dict(r) for r in reader]
    elif filename.lower().endswith((".xlsx", ".xls")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: (str(v) if v is not None else "") for i, v in enumerate(row)})
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to parse Excel: {e}")
    else:
        raise HTTPException(status_code=400, detail="Only .xlsx, .xls, or .csv files supported")

    # ── Column mapping (Excel OPL → schema) ───────────────────────────────────
    _COL_MAP = {
        "rele": "register", "relevance": "register",
        "input date": "input_date", "inputdate": "input_date",
        "input by": "input_by", "inputby": "input_by",
        "sw category": "sw_category", "swcategory": "sw_category",
        "topic": "topic", "sub topic": "sub_topic", "subtopic": "sub_topic",
        "description": "description",
        "remarks": "remarks", "remarks/status update": "remarks", "status update": "remarks",
        "priority": "priority",
        "status": "status",
        "risk": "risk_flag", "risk flag": "risk_flag",
        "risk/impact": "risk_impact", "risk impact": "risk_impact",
        "start date": "start_date", "startdate": "start_date",
        "end date": "due_date", "enddate": "due_date", "due date": "due_date",
        "pic": "responsible", "responsible": "responsible",
        "topic owner": "topic_owner", "topicowner": "topic_owner",
        "subject": "subject", "title": "subject",
        "category": "category", "source": "source",
        "owner": "owner", "type": "entry_type", "entry type": "entry_type",
        "sprint": "sprint_tag", "sprint tag": "sprint_tag",
        "tags": "tags",
    }

    existing = _load()
    existing_ids = {e["id"] for e in existing}
    today = _today()
    now = _now_iso()
    imported, skipped, errors = [], [], []

    for i, raw in enumerate(rows):
        mapped = {}
        for col, val in raw.items():
            key = col.strip().lower()
            field = _COL_MAP.get(key)
            if field:
                mapped[field] = val

        if not mapped.get("subject") and not mapped.get("topic") and not mapped.get("description"):
            skipped.append({"row": i + 2, "reason": "No subject/topic/description"})
            continue

        # Coerce types
        if "risk_flag" in mapped:
            mapped["risk_flag"] = str(mapped["risk_flag"]).lower() in ("true", "yes", "1", "x")
        if "responsible" in mapped and isinstance(mapped["responsible"], str):
            mapped["responsible"] = [r.strip() for r in mapped["responsible"].split(",") if r.strip()]
        if "tags" in mapped and isinstance(mapped["tags"], str):
            mapped["tags"] = [t.strip() for t in mapped["tags"].split(",") if t.strip()]

        # subject fallback
        if not mapped.get("subject"):
            mapped["subject"] = mapped.get("topic") or mapped.get("description", "")[:80]

        entry_id = mapped.get("id", "")
        if mode == "upsert" and entry_id and entry_id in existing_ids:
            idx = next(j for j, e in enumerate(existing) if e["id"] == entry_id)
            for k, v in mapped.items():
                if k not in {"id", "created_at", "input_date", "input_by"}:
                    existing[idx][k] = v
            existing[idx]["last_change"] = now
            imported.append({"row": i + 2, "id": entry_id, "action": "updated"})
        else:
            new_entry = {
                "entry_type": "Task", "parent_id": None,
                "subject": "", "description": "", "remarks": "", "last_note": "",
                "owner": "", "responsible": [], "information_to": [], "topic_owner": "",
                "sw_category": "", "topic": "", "sub_topic": "",
                "category": "", "source": "", "register": "",
                "tags": [], "sprint_tag": "", "meeting": "",
                "priority": "Medium", "status": "running",
                "risk_flag": False, "risk_impact": "",
                "confidential": False, "pinned": False,
                "input_date": today, "start_date": today,
                "due_date": (date.today() + timedelta(days=14)).isoformat(),
                "closed_date": None, "input_by": "",
                "notes": [], "attachments": [],
                "last_change": now, "last_change_by": "", "created_at": now,
            }
            new_entry.update(mapped)
            new_entry["id"] = _next_id(existing + imported)
            existing.append(new_entry)
            imported.append({"row": i + 2, "id": new_entry["id"], "action": "created"})

    _save(existing)
    return {
        "imported": len(imported),
        "skipped":  len(skipped),
        "errors":   errors,
        "details":  imported[:50],
    }


# ── GET /api/opl/export ───────────────────────────────────────────────────────

@router.get("/api/opl/export")
def export_opl(
    format:   str = Query("xlsx", description="xlsx | csv"),
    show:     str = Query("all"),
    category: str = Query(""),
    search:   str = Query(""),
):
    """Export current filtered view. Columns match original Excel OPL format."""
    # Reuse list logic inline
    entries = [_apply_overdue(e) for e in _load()]
    allowed = _SHOW_FILTERS.get(show)
    if allowed is not None:
        entries = [e for e in entries if e.get("status", "running") in allowed]
    if category:
        entries = [e for e in entries if e.get("category", "").lower() == category.lower()]
    if search:
        q = search.lower()
        entries = [e for e in entries if q in e.get("subject","").lower() or q in e.get("description","").lower()]

    # Column order matching original Excel OPL
    COLS = [
        ("id", "ID"), ("register", "Rele"),
        ("input_date", "Input date"), ("input_by", "Input by"),
        ("sw_category", "SW Category"), ("topic", "Topic"), ("sub_topic", "Sub topic"),
        ("subject", "Subject / Description"), ("description", "Description"),
        ("remarks", "Remarks/Status update"),
        ("priority", "Priority"), ("status", "Status"),
        ("risk_flag", "Risk"), ("risk_impact", "Risk/Impact"),
        ("start_date", "Start date"), ("due_date", "End date"),
        ("responsible", "PIC"), ("topic_owner", "Topic Owner"),
        ("entry_type", "Type"), ("category", "Category"), ("source", "Source"),
        ("owner", "Owner"), ("sprint_tag", "Sprint"), ("tags", "Tags"),
        ("last_change", "Last change"), ("last_note", "Last note"),
    ]

    def _cell(entry, field):
        v = entry.get(field, "")
        if isinstance(v, list):
            return ", ".join(str(x) for x in v)
        if isinstance(v, bool):
            return "Yes" if v else "No"
        return str(v) if v is not None else ""

    if format == "csv":
        import csv
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow([label for _, label in COLS])
        for e in entries:
            writer.writerow([_cell(e, f) for f, _ in COLS])
        buf.seek(0)
        return StreamingResponse(
            io.BytesIO(buf.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=super_opl_export.csv"},
        )

    # Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OPL Export"

    hdr_fill = PatternFill(fill_type="solid", fgColor="003366")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)

    for col_idx, (_, label) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    red_font   = Font(color="CC0000")
    today_str  = _today()

    for row_idx, e in enumerate(entries, 2):
        for col_idx, (field, _) in enumerate(COLS, 1):
            val = _cell(e, field)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=(field in ("description", "remarks", "last_note")))
            if field == "due_date" and val and val < today_str and e.get("status") not in ("closed","deleted","rejected"):
                cell.font = red_font

    # Column widths
    widths = {"ID": 10, "Rele": 10, "Input date": 12, "Input by": 14,
              "SW Category": 14, "Topic": 20, "Sub topic": 20,
              "Subject / Description": 40, "Description": 40,
              "Remarks/Status update": 35, "Priority": 10, "Status": 14,
              "Risk": 8, "Risk/Impact": 25, "Start date": 12, "End date": 12,
              "PIC": 20, "Topic Owner": 16, "Type": 14, "Category": 14,
              "Source": 16, "Owner": 14, "Sprint": 14, "Tags": 20,
              "Last change": 18, "Last note": 35}
    for col_idx, (_, label) in enumerate(COLS, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = widths.get(label, 16)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=super_opl_export.xlsx"},
    )
